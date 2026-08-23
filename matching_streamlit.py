# -*- coding: utf-8 -*-
"""
برنامج مطابقة الأسماء العربية - نسخة Streamlit
====================================================
نفس منطق matching_engine.py (التطبيع والمطابقة والتصنيف) بواجهة ويب
مبنية على Streamlit، تدعم:
  - مطابقة ملف هدف واحد مع قاعدة مرجعية واحدة.
  - مطابقة ملف هدف واحد مع قاعدتين مرجعيتين معًا (ودمج النتيجة بصف واحد).
  - عمود مدرسة اختياري (بكل ملف) يُستخدم لحسم حالات "أقارب" تلقائيًا.
  - أعمدة إضافية اختيارية من القاعدة المرجعية (بالوضع المفرد) تُجلب تلقائيًا.
  - تنزيل ملف النتيجة الكامل وملف الملخص كإكسل منسّق ومُلوّن.

تشغيل:
    pip install streamlit pandas openpyxl
    streamlit run matching_streamlit.py
"""
import io
import datetime

import pandas as pd
import streamlit as st

from matching_engine import (
    run_matching, summarize, CATEGORY_ORDER, CATEGORY_MEANING, CATEGORY_COLOR,
    run_matching_dual, summarize_dual, DUAL_KIND_ORDER, DUAL_KIND_COLOR,
)

APP_TITLE = "برنامج مطابقة الأسماء العربية"
APP_VERSION = "1.0 (Streamlit)"
FONT_NAME = "Arial"

DUAL_KIND_MEANING = {
    'both': "الاسم موجود ومؤكد بكلتا القاعدتين المرجعيتين.",
    'only1': "الاسم موجود ومؤكد بالقاعدة الأولى فقط، ولا يوجد له سجل بالثانية.",
    'only2': "الاسم موجود ومؤكد بالقاعدة الثانية فقط، ولا يوجد له سجل بالأولى.",
    'review': "يوجد تشابه أو تقارب بأي من القاعدتين (وليس تطابقًا تامًا) - يحتاج فرز يدوي.",
    'not_found': "لا يوجد أي سجل مطابق له بأي من القاعدتين.",
}


# ============================================================================
# قراءة ملفات الإكسل (مع تخزين مؤقت لتجنب إعادة القراءة بكل إعادة تشغيل)
# ============================================================================

@st.cache_data(show_spinner=False)
def cached_sheet_names(file_bytes):
    xl = pd.ExcelFile(io.BytesIO(file_bytes))
    return xl.sheet_names


@st.cache_data(show_spinner=False)
def cached_columns(file_bytes, sheet):
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, nrows=5)
    return list(df.columns)


@st.cache_data(show_spinner=False)
def cached_names_extra_and_school(file_bytes, sheet, col, extra_cols, school_col):
    """يرجع (أسماء، dict أعمدة إضافية، قائمة المدرسة أو None) بنفس ترتيب/تصفية
    عمود الأسماء الرئيسي، بحيث تبقى فهارس ref_indices ومحاذاة المدرسة صالحة."""
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet)
    mask = df[col].notna()
    filtered = df.loc[mask]
    names = filtered[col].astype(str).tolist()
    extra_data = {c: filtered[c].fillna('').astype(str).tolist() for c in extra_cols}
    school = filtered[school_col].fillna('').astype(str).tolist() if school_col else None
    return names, extra_data, school


# ============================================================================
# كتابة ملفات الإكسل الناتجة (نفس تنسيق نسخة سطح المكتب حرفيًا)
# ============================================================================

def build_output_workbooks(rows, target_col_label, extra_col_names=None, extra_col_values=None):
    """يبني ملفي (النتيجة، الملخص) كإكسل ويرجع كائني Workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    extra_col_names = extra_col_names or []
    extra_col_values = extra_col_values or [{} for _ in rows]

    header_fill = PatternFill('solid', fgColor='2F5496')
    header_font = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=11)
    title_font = Font(name=FONT_NAME, bold=True, size=14, color='2F5496')
    sub_font = Font(name=FONT_NAME, size=10, italic=True, color='595959')
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    counts = summarize(rows)
    total = len(rows)

    order_index = {c: i for i, c in enumerate(CATEGORY_ORDER)}
    ordered = sorted(range(len(rows)), key=lambda i: (order_index.get(rows[i].category, 99), i))

    # ------------------------------------------------------------------
    # الملف الرئيسي
    # ------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "نتيجة المطابقة"
    ws.sheet_view.rightToLeft = True

    cols = ['ت', target_col_label, 'الاسم المطابق / المرشحون بقاعدة العقود',
            *extra_col_names, 'نسبة التطابق %', 'الفئة', 'التفسير']
    n_cols = len(cols)
    last_col_letter = get_column_letter(n_cols)

    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = "نتيجة مطابقة الأسماء - نسخة كاملة مصفّاة ومبوّبة"
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 24

    summary_line = " | ".join(f"{c}: {counts.get(c,0)}" for c in CATEGORY_ORDER)
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = f"الإجمالي: {total} | " + summary_line
    ws['A2'].font = sub_font

    header_row = 4
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=j, value=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    r = header_row + 1
    for i in ordered:
        row = rows[i]
        fill_color = CATEGORY_COLOR.get(row.category)
        extra_vals = [extra_col_values[i].get(c, '') for c in extra_col_names]
        vals = [row.idx + 1, row.original_name, row.match_text, *extra_vals, row.score,
                row.category, row.note]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            cell.border = border
            if fill_color:
                cell.fill = PatternFill('solid', fgColor=fill_color)
        r += 1

    widths = [7, 34, 34, *([26] * len(extra_col_names)), 14, 20, 55]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f'A{header_row+1}'
    ws.auto_filter.ref = f'A{header_row}:{last_col_letter}{r-1}'

    # ------------------------------------------------------------------
    # ملف الملخص
    # ------------------------------------------------------------------
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "ملخص النتيجة"
    ws2.sheet_view.rightToLeft = True

    ws2.merge_cells('A1:C1')
    ws2['A1'] = "ملخص نتيجة المطابقة"
    ws2['A1'].font = title_font
    ws2.row_dimensions[1].height = 26

    ws2.merge_cells('A2:C2')
    ws2['A2'] = f"الإجمالي: {total} اسم"
    ws2['A2'].font = Font(name=FONT_NAME, size=11, italic=True, color='595959')

    headers = ['الفئة', 'العدد', 'يعني إيش؟']
    hr = 4
    for j, h in enumerate(headers, start=1):
        cell = ws2.cell(row=hr, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    rr = hr + 1
    for cat in CATEGORY_ORDER:
        fill = PatternFill('solid', fgColor=CATEGORY_COLOR[cat])
        vals = [cat, counts.get(cat, 0), CATEGORY_MEANING[cat]]
        for j, v in enumerate(vals, start=1):
            cell = ws2.cell(row=rr, column=j, value=v)
            cell.font = Font(name=FONT_NAME, size=11, bold=(j == 1))
            cell.alignment = Alignment(horizontal='right' if j != 2 else 'center',
                                        vertical='center', wrap_text=True)
            cell.border = border
            cell.fill = fill
        ws2.row_dimensions[rr].height = 32
        rr += 1

    ws2.cell(row=rr, column=1, value='الإجمالي').font = Font(name=FONT_NAME, bold=True, size=11)
    ws2.cell(row=rr, column=2, value=total).font = Font(name=FONT_NAME, bold=True, size=11)
    for c in (1, 2, 3):
        ws2.cell(row=rr, column=c).border = border

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 70

    return wb, wb2


def build_dual_output_workbooks(rows, target_col_label, label1, label2):
    """يبني ملفي (النتيجة، الملخص) لمقارنة قاعدتين مرجعيتين معًا ويرجع كائني Workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill('solid', fgColor='2F5496')
    header_font = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=11)
    title_font = Font(name=FONT_NAME, bold=True, size=14, color='2F5496')
    sub_font = Font(name=FONT_NAME, size=10, italic=True, color='595959')
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    counts = summarize_dual(rows)
    total = len(rows)
    kind_display = {
        'both': "مؤكد بالقاعدتين", 'only1': f"مؤكد بـ{label1} فقط",
        'only2': f"مؤكد بـ{label2} فقط", 'review': "متشابه/قريب - يحتاج مراجعة",
        'not_found': "غير موجود بالقاعدتين",
    }

    order_index = {k: i for i, k in enumerate(DUAL_KIND_ORDER)}
    ordered = sorted(range(len(rows)), key=lambda i: (order_index.get(rows[i].status_kind, 99), i))

    # ------------------------------------------------------------------
    # الملف الرئيسي
    # ------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "مقارنة القاعدتين"
    ws.sheet_view.rightToLeft = True

    cols = ['ت', target_col_label, f'مطابق {label1}', f'فئة {label1}', f'نسبة % {label1}',
            f'مطابق {label2}', f'فئة {label2}', f'نسبة % {label2}', 'الحالة المدمجة', 'التفسير']
    n_cols = len(cols)
    last_col_letter = get_column_letter(n_cols)

    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = f"مقارنة الأسماء بين قاعدتي {label1} و{label2}"
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 24

    summary_line = " | ".join(f"{kind_display[k]}: {counts.get(k,0)}" for k in DUAL_KIND_ORDER)
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = f"الإجمالي: {total} | " + summary_line
    ws['A2'].font = sub_font

    header_row = 4
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=j, value=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    r = header_row + 1
    for i in ordered:
        row = rows[i]
        fill_color = DUAL_KIND_COLOR.get(row.status_kind)
        vals = [row.idx + 1, row.original_name,
                row.row1.match_text, row.row1.category, row.row1.score,
                row.row2.match_text, row.row2.category, row.row2.score,
                row.status_label, row.note]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            cell.border = border
            if fill_color:
                cell.fill = PatternFill('solid', fgColor=fill_color)
        r += 1

    widths = [7, 30, 28, 16, 12, 28, 16, 12, 22, 45]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f'A{header_row+1}'
    ws.auto_filter.ref = f'A{header_row}:{last_col_letter}{r-1}'

    # ------------------------------------------------------------------
    # ملف الملخص
    # ------------------------------------------------------------------
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "ملخص المقارنة"
    ws2.sheet_view.rightToLeft = True

    ws2.merge_cells('A1:C1')
    ws2['A1'] = "ملخص مقارنة القاعدتين"
    ws2['A1'].font = title_font
    ws2.row_dimensions[1].height = 26

    ws2.merge_cells('A2:C2')
    ws2['A2'] = f"الإجمالي: {total} اسم"
    ws2['A2'].font = Font(name=FONT_NAME, size=11, italic=True, color='595959')

    headers = ['الحالة', 'العدد', 'يعني إيش؟']
    hr = 4
    for j, h in enumerate(headers, start=1):
        cell = ws2.cell(row=hr, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    rr = hr + 1
    for k in DUAL_KIND_ORDER:
        fill = PatternFill('solid', fgColor=DUAL_KIND_COLOR[k])
        vals = [kind_display[k], counts.get(k, 0), DUAL_KIND_MEANING[k]]
        for j, v in enumerate(vals, start=1):
            cell = ws2.cell(row=rr, column=j, value=v)
            cell.font = Font(name=FONT_NAME, size=11, bold=(j == 1))
            cell.alignment = Alignment(horizontal='right' if j != 2 else 'center',
                                        vertical='center', wrap_text=True)
            cell.border = border
            cell.fill = fill
        ws2.row_dimensions[rr].height = 32
        rr += 1

    ws2.cell(row=rr, column=1, value='الإجمالي').font = Font(name=FONT_NAME, bold=True, size=11)
    ws2.cell(row=rr, column=2, value=total).font = Font(name=FONT_NAME, bold=True, size=11)
    for c in (1, 2, 3):
        ws2.cell(row=rr, column=c).border = border

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 70

    return wb, wb2


def workbook_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================================
# واجهة اختيار ملف + ورقة + عمود (+ أعمدة إضافية اختياريًا)
# ============================================================================

def guess_default_col(cols):
    for i, c in enumerate(cols):
        if "اسم" in str(c):
            return i
    return 0


def render_file_box(key_prefix, title, allow_extra_cols=False,
                     show_label_entry=False, default_label="القاعدة"):
    """يعرض صندوق اختيار ملف/ورقة/عمود ويرجع dict بالإعدادات أو None إن لم يكتمل."""
    with st.container(border=True):
        st.markdown(f"##### {title}")

        label_val = default_label
        if show_label_entry:
            label_val = st.text_input(
                "اسم القاعدة (يظهر بتقرير المقارنة):",
                value=default_label, key=f"{key_prefix}_label")

        uploaded = st.file_uploader(
            "اختر ملف إكسل...", type=["xlsx", "xls"], key=f"{key_prefix}_file")

        if uploaded is None:
            return None

        file_bytes = uploaded.getvalue()
        try:
            sheets = cached_sheet_names(file_bytes)
        except Exception as e:
            st.error(f"تعذّر فتح الملف: {e}")
            return None

        c1, c2 = st.columns(2)
        with c1:
            sheet = st.selectbox("الورقة:", sheets, key=f"{key_prefix}_sheet")
        try:
            cols = cached_columns(file_bytes, sheet)
        except Exception as e:
            st.error(f"تعذّر قراءة الورقة: {e}")
            return None

        with c2:
            col = st.selectbox("عمود الأسماء:", cols,
                                index=guess_default_col(cols), key=f"{key_prefix}_col")

        school_options = ["(بدون)"] + cols
        school_sel = st.selectbox(
            "عمود المدرسة (اختياري - للتأكيد فقط):",
            school_options, key=f"{key_prefix}_school")
        school_col = school_sel if school_sel != "(بدون)" else None

        extra_cols = []
        if allow_extra_cols:
            extra_choices = [c for c in cols if c != col]
            extra_cols = st.multiselect(
                "أعمدة إضافية تُجلب تلقائيًا عند المطابقة (اختياري):",
                extra_choices, key=f"{key_prefix}_extra")

        return {
            "file_bytes": file_bytes,
            "filename": uploaded.name,
            "sheet": sheet,
            "col": col,
            "school_col": school_col,
            "extra_cols": extra_cols,
            "label": (label_val.strip() or default_label) if show_label_entry else default_label,
        }


# ============================================================================
# عرض الجداول الملوّنة
# ============================================================================

def show_single_results(rows, target_col_label):
    counts = summarize(rows)
    total = len(rows)

    st.markdown("#### ملخص النتيجة")
    cols_widgets = st.columns(len(CATEGORY_ORDER))
    for w, cat in zip(cols_widgets, CATEGORY_ORDER):
        with w:
            st.markdown(
                f"""<div style="background:#{CATEGORY_COLOR[cat]};border-radius:8px;
                padding:10px;text-align:center;">
                <div style="font-size:22px;font-weight:bold;">{counts.get(cat,0)}</div>
                <div style="font-size:12px;">{cat}</div>
                </div>""", unsafe_allow_html=True)
    st.caption(f"الإجمالي: {total} اسم")

    with st.expander("يعني إيش كل فئة؟"):
        for cat in CATEGORY_ORDER:
            st.markdown(f"**{cat}**: {CATEGORY_MEANING[cat]}")

    order_index = {c: i for i, c in enumerate(CATEGORY_ORDER)}
    ordered = sorted(range(len(rows)), key=lambda i: (order_index.get(rows[i].category, 99), i))

    st.markdown("#### معاينة النتائج")
    filt = st.multiselect("تصفية حسب الفئة:", CATEGORY_ORDER, default=[], key="filter_single")
    show_rows = [i for i in ordered if not filt or rows[i].category in filt]

    data = {
        "ت": [rows[i].idx + 1 for i in show_rows],
        target_col_label: [rows[i].original_name for i in show_rows],
        "الاسم المطابق / المرشحون": [rows[i].match_text for i in show_rows],
        "نسبة %": [rows[i].score for i in show_rows],
        "الفئة": [rows[i].category for i in show_rows],
        "التفسير": [rows[i].note for i in show_rows],
    }
    df = pd.DataFrame(data)
    cat_list = [rows[i].category for i in show_rows]

    def highlight(row):
        color = CATEGORY_COLOR.get(cat_list[row.name])
        return [f'background-color: #{color}' if color else '' for _ in row]

    st.dataframe(df.style.apply(highlight, axis=1), use_container_width=True, height=520)


def show_dual_results(rows, target_col_label, label1, label2):
    counts = summarize_dual(rows)
    total = len(rows)
    kind_display = {
        'both': "مؤكد بالقاعدتين",
        'only1': f"مؤكد بـ{label1} فقط",
        'only2': f"مؤكد بـ{label2} فقط",
        'review': "متشابه/قريب - يحتاج مراجعة",
        'not_found': "غير موجود بالقاعدتين",
    }

    st.markdown("#### ملخص المقارنة")
    cols_widgets = st.columns(len(DUAL_KIND_ORDER))
    for w, k in zip(cols_widgets, DUAL_KIND_ORDER):
        with w:
            st.markdown(
                f"""<div style="background:#{DUAL_KIND_COLOR[k]};border-radius:8px;
                padding:10px;text-align:center;">
                <div style="font-size:22px;font-weight:bold;">{counts.get(k,0)}</div>
                <div style="font-size:12px;">{kind_display[k]}</div>
                </div>""", unsafe_allow_html=True)
    st.caption(f"الإجمالي: {total} اسم")

    with st.expander("يعني إيش كل حالة؟"):
        for k in DUAL_KIND_ORDER:
            st.markdown(f"**{kind_display[k]}**: {DUAL_KIND_MEANING[k]}")

    order_index = {k: i for i, k in enumerate(DUAL_KIND_ORDER)}
    ordered = sorted(range(len(rows)), key=lambda i: (order_index.get(rows[i].status_kind, 99), i))

    st.markdown("#### معاينة النتائج")
    filt = st.multiselect(
        "تصفية حسب الحالة:",
        [kind_display[k] for k in DUAL_KIND_ORDER], default=[], key="filter_dual")
    filt_kinds = {k for k in DUAL_KIND_ORDER if kind_display[k] in filt}
    show_rows = [i for i in ordered if not filt_kinds or rows[i].status_kind in filt_kinds]

    data = {
        "ت": [rows[i].idx + 1 for i in show_rows],
        target_col_label: [rows[i].original_name for i in show_rows],
        f"مطابق {label1}": [rows[i].row1.match_text for i in show_rows],
        f"فئة {label1}": [rows[i].row1.category for i in show_rows],
        f"مطابق {label2}": [rows[i].row2.match_text for i in show_rows],
        f"فئة {label2}": [rows[i].row2.category for i in show_rows],
        "الحالة المدمجة": [rows[i].status_label for i in show_rows],
        "التفسير": [rows[i].note for i in show_rows],
    }
    df = pd.DataFrame(data)
    kind_list = [rows[i].status_kind for i in show_rows]

    def highlight(row):
        color = DUAL_KIND_COLOR.get(kind_list[row.name])
        return [f'background-color: #{color}' if color else '' for _ in row]

    st.dataframe(df.style.apply(highlight, axis=1), use_container_width=True, height=520)


# ============================================================================
# التطبيق الرئيسي
# ============================================================================

def main():
    st.set_page_config(page_title=APP_TITLE, page_icon="🔎", layout="wide")

    st.markdown(
        """
        <style>
        html, body, [class*="css"]  { direction: rtl; }
        .stMarkdown, .stDataFrame, .stTextInput, .stSelectbox, .stMultiSelect,
        .stFileUploader, .stButton, .stCaption, .stExpander { direction: rtl; text-align: right; }
        div[data-testid="stMetricValue"] { direction: rtl; }
        thead tr th { text-align: right !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.title(APP_TITLE)
    st.caption("يقارن ملف أسماء (مثل ملف مستبعدين) مع قاعدة بيانات مرجعية (مثل قاعدة عقود)")
    st.caption(f"إصدار {APP_VERSION}")
    st.divider()

    for key in ("result_data",):
        if key not in st.session_state:
            st.session_state[key] = None

    col_a, col_b = st.columns(2)
    with col_a:
        target_box = render_file_box("target", "ملف الأسماء (الهدف)")
    with col_b:
        ref_box = render_file_box(
            "ref1", "ملف القاعدة المرجعية الأولى", allow_extra_cols=True,
            show_label_entry=True, default_label="القاعدة الأولى")

    dual_mode = st.checkbox(
        "مقارنة مع قاعدة مرجعية ثانية أيضًا (لتمييز الأسماء المشتركة/المتقاربة بين القاعدتين)")

    ref2_box = None
    if dual_mode:
        ref2_box = render_file_box(
            "ref2", "ملف القاعدة المرجعية الثانية",
            show_label_entry=True, default_label="القاعدة الثانية")

    st.divider()

    ready = bool(target_box and ref_box and (not dual_mode or ref2_box))
    run_clicked = st.button("ابدأ المطابقة", type="primary", disabled=not ready)

    if not ready:
        st.info("الرجاء اختيار الملفات المطلوبة وتحديد عمود الأسماء بكل منها للمتابعة.")

    if run_clicked and ready:
        progress_bar = st.progress(0, text="جارٍ التحميل...")

        def progress_cb(pct, msg):
            progress_bar.progress(min(max(int(pct), 0), 100), text=msg)

        try:
            target_names, _, target_school = cached_names_extra_and_school(
                target_box["file_bytes"], target_box["sheet"], target_box["col"],
                (), target_box["school_col"])

            if not target_names:
                raise ValueError("عمود أسماء الهدف فارغ - تأكد من اختيار العمود الصحيح.")

            if dual_mode:
                ref1_names, _, ref1_school = cached_names_extra_and_school(
                    ref_box["file_bytes"], ref_box["sheet"], ref_box["col"],
                    (), ref_box["school_col"])
                ref2_names, _, ref2_school = cached_names_extra_and_school(
                    ref2_box["file_bytes"], ref2_box["sheet"], ref2_box["col"],
                    (), ref2_box["school_col"])

                if not ref1_names or not ref2_names:
                    raise ValueError("عمود أسماء إحدى القاعدتين المرجعيتين فارغ - تأكد من اختيار العمود الصحيح.")

                rows = run_matching_dual(
                    target_names, ref1_names, ref_box["label"], ref2_names, ref2_box["label"],
                    progress_cb=progress_cb, target_school=target_school,
                    ref1_school=ref1_school, ref2_school=ref2_school)

                progress_bar.progress(95, text="جارٍ بناء ملفات الإكسل...")
                wb, wb2 = build_dual_output_workbooks(
                    rows, target_box["col"], ref_box["label"], ref2_box["label"])
                counts = summarize_dual(rows)
                base_name = target_box["filename"].rsplit('.', 1)[0]

                st.session_state["result_data"] = {
                    "mode": "dual",
                    "rows": rows,
                    "target_col": target_box["col"],
                    "label1": ref_box["label"],
                    "label2": ref2_box["label"],
                    "base_name": base_name,
                    "wb_bytes": workbook_to_bytes(wb),
                    "wb2_bytes": workbook_to_bytes(wb2),
                    "counts": counts,
                    "total": len(target_names),
                }
            else:
                ref_names, ref_extra_data, ref_school = cached_names_extra_and_school(
                    ref_box["file_bytes"], ref_box["sheet"], ref_box["col"],
                    tuple(ref_box["extra_cols"]), ref_box["school_col"])

                if not ref_names:
                    raise ValueError("عمود أسماء القاعدة المرجعية فارغ - تأكد من اختيار العمود الصحيح.")

                rows = run_matching(
                    target_names, ref_names, progress_cb=progress_cb,
                    target_school=target_school, ref_school=ref_school)

                progress_bar.progress(90, text="جارٍ تجهيز الأعمدة الإضافية...")
                extra_col_values = []
                for row in rows:
                    vals = {}
                    for c in ref_box["extra_cols"]:
                        col_data = ref_extra_data.get(c, [])
                        vals[c] = " | ".join(
                            col_data[ri] for ri in row.ref_indices if 0 <= ri < len(col_data))
                    extra_col_values.append(vals)

                progress_bar.progress(95, text="جارٍ بناء ملفات الإكسل...")
                wb, wb2 = build_output_workbooks(
                    rows, target_box["col"], extra_col_names=ref_box["extra_cols"],
                    extra_col_values=extra_col_values)
                counts = summarize(rows)
                base_name = target_box["filename"].rsplit('.', 1)[0]

                st.session_state["result_data"] = {
                    "mode": "single",
                    "rows": rows,
                    "target_col": target_box["col"],
                    "base_name": base_name,
                    "wb_bytes": workbook_to_bytes(wb),
                    "wb2_bytes": workbook_to_bytes(wb2),
                    "counts": counts,
                    "total": len(target_names),
                }

            progress_bar.progress(100, text="اكتملت المطابقة")
            st.success(f"تمت مطابقة {st.session_state['result_data']['total']} اسمًا بنجاح.")

        except Exception as e:
            st.error(f"حدث خطأ: {e}")
            st.session_state["result_data"] = None

    result = st.session_state.get("result_data")
    if result:
        st.divider()
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")

        dl1, dl2 = st.columns(2)
        with dl1:
            st.download_button(
                "⬇ حفظ ملف النتيجة",
                data=result["wb_bytes"],
                file_name=f"{result['base_name']}_نتيجة_المطابقة_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with dl2:
            st.download_button(
                "⬇ حفظ ملف الملخص",
                data=result["wb2_bytes"],
                file_name=f"{result['base_name']}_ملخص_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        st.divider()
        if result["mode"] == "dual":
            show_dual_results(result["rows"], result["target_col"], result["label1"], result["label2"])
        else:
            show_single_results(result["rows"], result["target_col"])


if __name__ == "__main__":
    main()
